
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
#wb=Workbook()
#wb.save('Habitats.xlsx') 
dh={}
for i in range (1, 10):
    url="https://pokeapi.co/api/v2/pokemon-habitat/"+str(i)+"/"
    response = requests.request("GET", url)
    data=response.json()
    k=(data['name'])
    nam=k
    dh[nam]=[]
    j=0
    try:
        for j in range(50000000000000000000000000000000):
            ko=(data['pokemon_species'][j]['name'])
            dh[nam].append(ko)
            j+1
    except IndexError:
        pass
#cambiar longitud antes de agregarlo
max_length = max(len(v) for v in dh.values())
for key, values in dh.items():
    while len(values) < max_length:
        values.append(0)
#ESCRIBIENDO EN EL WB
pat = "Habitats.xlsx"  
wb = load_workbook(pat)
ws = wb.active
start_row = ws.max_row + 1
a=list(dh.keys())
ws.append(a)
start_row += 1
for row in zip(*dh.values()):
    ws.append(row)

wb.save(pat)
print('Se ha guardado en el excel '+str(pat))
