import os

## SOLO CORRER DESDE MAIN.PY

def estadisticas():
    path=os.getcwd()
    pathModulos=os.path.join(path,'Modulos')
    pathGraficas=os.path.join(pathModulos,'Graficas')
    def ejecutarmenu(): 
        while True:
            print("\nSeleccione una opción:\n1. Cantidad de Pokémon por color\n2. El color que tiene más Pokémon\n3. Cantidad de Pokémon por su grupo de huevo\n4. El grupo de huevo con más Pokémon\n5. La cantidad de cada hábitat\n6. Salirse")
            r1=str(input(">> "))
            if r1=='1':
                cua_cada(wb,ws,dic)
            elif r1=='2':
                printmodaco(ws,ind,ma)
            elif r1=='3':
                cgp(pagina3,nongh)
            elif r1=='4':
                printgpp(nongh,ma,ngh,pagina3)
            elif r1=='5':
                pathFile=os.path.join(pathGraficas, 'Habitats.xlsx')
                wb2=load_workbook(filename=pathFile)
                ws2=wb2.active
                porch(ws2)
            elif r1=='6':
                print('')
                print('')
                print('')
                return
            else:
                pass
                print("Ingrese una de las opciones brindadas")
            print('')
    def modaco(wb,ws):
        s=-1
        listas={}
        for i in range(0,10):
            c=i+1
            s=s+1
            nam=f"L{c}"
            listas[nam]=[]
            for row in ws.iter_rows(min_row=2, max_col=c, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    listas[nam].append(valor)
                    continue
        dic=[]
        for i in range (0,10):
            d=len(listas['L'+str(i+1)])
            dic.append(d)
        ma=0
        for i in range (0,len(dic)):
            if dic[i]>ma:
                ma=dic[i]  
            else:
                pass
        ind=dic.index(ma)
        key=list(listas.keys())
        key[ind]
        wb.save(os.path.join(pathGraficas,"GrColor.xlsx"))
        return ind, ma, dic

    def printmodaco(ws,ind,ma):
        m=ws.cell(row=1, column=(ind+1))
        print("\nEl color que más se repite es " +str(m.value)+ " con " + str(ma) +" Pokémon")
        exit
        
    def cua_cada(wb,ws,dic):
        while True:
                k=str(input("\nSelecciona una opción \n1.Todos los colores\n2.Uno en específico\n3.Volver al menú\n>> "))
                if k.isdigit():
                    k=int(k)
                else:
                    k=0
                if k==1:
                    for i in range (1,11):
                        cv=ws.cell(row=1,column=i)
                        va=dic[i-1]
                        print("\nEl color "+ str(cv.value) +" tiene " + str(va)+" Pokémon")
                elif k==2:
                    cuacaop2(wb,ws,dic)

                elif k==3:
                    print("Volviendo al menú....")
                    break
                else:
                    print("Ingrese nuevamente\n")
    def cuacaop2(wb,ws,dic):
        ola=[]
        diccolores={}
        for j in range (1,11):
            m=ws.cell(row=1, column=(j))
            m=m.value
            ola.append(m)
            diccolores[j]=ola[j-1]
        while True:
                print("\nIngrese un opción de acuerdo al índice que corresponda")
                for i in range(1,11):
                      print(str(i)+': '+diccolores[i])
                print("11. Salirse")
                try:
                    n=(int(input(">> ")))
                except:
                    print('Vuelva a intentar')
                    continue
                if n==11:
                    return
                elif n<11 and n>0:
                    va=dic[n-1]
                    print("\nEl color "+str(diccolores[n])+" contiene "+str(va)+" Pokémon")
                else:
                    print("Ingrese de nuevo")

    def porch(ws2):
        #Porcentaje de cada habitat
        s=-1
        hab={}
        for i in range(0,9):
            c=i+1
            s=s+1
            nam=f"L{c}"
            hab[nam]=[]
            for row in ws2.iter_rows(min_row=2, max_col=c, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    hab[nam].append(valor)
                    continue
        ongdh=[]
        for i in range (0,9):
            d=len(hab['L'+str(i+1)])
            ongdh.append(d)
        s=sum(ongdh)
        por=[]
        for i in range (0,9):
            a=ongdh[i]
            c=a*(100/s)
            por.append(a)
        sn=1
        while True:
            print("\nIngrese el número del hábitat del que desea conocer su porcentaje:\n1. Cueva\n2. Bosque\n3. Campo\n4. Montaña\n5. Raro\n6. Terreno rocoso\n7. Océano\n8. Urbano\n9. Costas\n10. Volver al menú principal")
            r=str(input(">> "))
            if r.isdigit():
                r=int(r)
            else:
                r=0
            if 1<=r<=9:
                for i in range (1,10):
                    if r==i:
                        t=ws2.cell(row=1, column=i)
                        x=t.value
                        y=x.capitalize()
                        print('\nHay '+str(por[i-1])+' del hábitat '+y)
            elif r==10:
                break

            else:
                print("Ingrese de nuevo")
    def cgp(ws3,nongh):
        while True:
            r=int(input("\nSeleccione una opción:\n1.Vizualizar la cantidad de todos los grupos de huevos\n2.Visualizar un cierto tipo\n3.Volver al menú\n>> "))
            if r==1:
                for i in range (1,15):
                    cv=ws3.cell(row=1,column=i)
                    cv=cv.value
                    cv=cv.capitalize()
                    va=nongh[i-1]
                    print("\nEl grupo de huevo "+ str(cv) +" tiene " + str(va)+" Pokémon\n")
            elif r==2:
                pricgp(pagina3,nongh)
            elif r==3:
                return
            else:
                print("Ingrese nuevamente")
    def pricgp(ws3,nongh):
        titulos=[]
        opc={}
        for i in range (1,15):
            m=ws3.cell(row=1, column=(i))
            m=m.value
            titulos.append(m)
            opc[i]=titulos[i-1]
        while True:
            print("\nIngrese una opción")
            for i in range(1,15):
                  print(str(i)+': '+opc[i])
            print('15.Volver atrás')
            u=str(input('>> '))
            if u.isdigit():
                u=int(u)
            else:
                u=0
            if u==15:
                return
            elif u<15 and u>0:
                va=nongh[u-1]
                print("\nEl grupo de huevo "+str(opc[u])+" contiene "+str(va)+" Pokémon")
            else:
                print("Intente de nuevo\n")
    def gpp(pagina3):
        s=-1
        ngh={}
        for i in range(0,16):
            c=i+1
            s=s+1
            nam=f"L{c}"
            ngh[nam]=[]
            for row in pagina3.iter_rows(min_row=2, max_col=c, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    ngh[nam].append(valor)
                    continue
        nongh=[]
        for i in range (0,16):
            d=len(ngh['L'+str(i+1)])
            nongh.append(d)
        ma=0
        for i in range (0,len(nongh)):
            if nongh[i]>ma:
                ma=nongh[i]  
            else:
                pass
        worbe.save(os.path.join(pathGraficas,"Gpohvs.xlsx"))
        return nongh,ma,ngh,pagina3
    def printgpp(nongh,ma,ngh,pagina3):
        ind=nongh.index(ma)
        key=list(ngh.keys())
        key[ind]
        m=pagina3.cell(row=1, column=(ind+1))
        m=m.value
        m=m.capitalize()
        print("\nEl grupo de huevo que más se repite es " +str(m)+ " con " + str(ma) +" Pokémon")
    from openpyxl import load_workbook
    from openpyxl import Workbook
    import difflib
    pathFile=os.path.join(pathGraficas, 'GrColor.xlsx')
    wb =load_workbook(filename=pathFile)
    ws = wb.active
    ind, ma,dic=modaco(wb,ws)
    pathFile=os.path.join(pathGraficas, 'Gpohvs.xlsx')
    worbe=load_workbook(filename=pathFile)
    pagina3=worbe.active
    nongh,ma,ngh,pagina3=gpp(pagina3)
    ejecutarmenu()
    return
