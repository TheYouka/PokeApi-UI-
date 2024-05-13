def agregarcolores():
    LCN, LCB, LCC, LCG, LCV, LCR, LCM, LCRO, LCBLA, LCAM= [], [], [], [], [], [], [], [], [], []
    for t in range(10):
        url = "https://pokeapi.co/api/v2/pokemon-color/"+str(t+1)+"/"
        u=t+1
        payload = {}
        headers = {}
        response = requests.request("GET", url)
        data=response.json()
        if u==1:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCN.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==2:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCB.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==3:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCC.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==4:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCG.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==5:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCV.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==6:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCR.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==7:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCM.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==8:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCRO.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==9:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCBLA.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==10:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    LCAM.append(ko)
                    i+1
            except IndexError:
                continue
        else:
            exit
        
    x={'Negro':LCN, 'Azul':LCB, 'Café':LCC, 'Gris':LCG, 'Verde':LCV, 'Rosa':LCR, 'Morado':LCM, 'Rojo':LCRO, 'Blanco':LCBLA, 'Amarillo':LCAM}
    return x

import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook

dicb=agregarcolores()

#LONGITUD
max_length = max(len(v) for v in dicb.values())
for key, values in dicb.items():
    while len(values) < max_length:
        values.append(0)
#CREANDO EL WB
#wb=Workbook()
#wb.save('GrColor.xlsx')
#ESCRIBIENDO EN EL WB
pat = "GrColor.xlsx"  
wb = load_workbook(pat)
ws = wb.active
start_row = ws.max_row + 1
a=list(dicb.keys())
ws.append(a)
start_row += 1
for row in zip(*dicb.values()):
    ws.append(row)

wb.save(pat)
#print('Se ha guardado en el excel '+str(pat)')
