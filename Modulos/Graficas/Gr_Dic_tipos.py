from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import os

#grafica de pastel de tipos de pokemones
def grafTipos(fila=1,archivo='Gráficas.xlsx'):
    try:
        
        path=os.getcwd()
        pathConsultas=os.path.join(path,'Consultas')
        pathModulos=os.path.join(path,'Modulos')
        pathGraficas=os.path.join(pathModulos,'Graficas')
        pathExcel=pathGraficas=os.path.join(pathGraficas,'GrType.xlsx')
        archivoFinal=os.path.join(pathConsultas,archivo)
        
        wb =load_workbook(pathExcel)
        ws = wb.active
        s=-1
        listas={}
        for i in range(0,18):
            c=i+1
            s=s+1
            nam=f"L{c}"
            listas[nam]=[]
            for row in ws.iter_rows(min_row=2, max_col=c, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    listas[nam].append(valor)
                    continue
        #Sacar la long de c/u
        tl=[]
        for i in range (0,18):
            d=len(listas['L'+str(i+1)])
            tl.append(d)
        #sacar total y porcentajes de c/u
        total=sum(tl)
        porcentajes_tip=[]
        for i in range (18):
            a=tl[i]
            p=(a*100)/1025
            porcentajes_tip.append(p)
        #Para asignar los titulos en una lista
        titulos=[]
        for i in range (1,19):
            t=ws.cell(row=1, column=i)
            titulos.append(t.value)

        #Crear gráfica
        plt.pie(porcentajes_tip, labels=titulos, autopct="%1.1f%%", startangle=90)
        plt.axis("equal")
        plt.title("Cantidad de Pokémon de cada tipo")

        wb.save(pathExcel)

        #Guardar la gráfica como imagen (esto es temporal)
        image_path = os.path.join(pathConsultas, 'grafica.png')
        plt.savefig(image_path)

        #se muestra la gráfica al usuario
        plt.show()

        
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Tipos'
        if 'Tipos' in wb.sheetnames:
            ws = wb['Tipos']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Tipos')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)
        
        #Eliminar la imagen temporal
        os.remove(image_path)

        return True

    except Exception as e:
        print(f"Error: {e}")
        return False
