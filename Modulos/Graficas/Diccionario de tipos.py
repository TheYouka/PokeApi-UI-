def tipos():
    L1, L2, L3, L4, L5, L6, L7, L8, L9, L10, L11, L12, L13, L14, L15, L16, L17, L18= [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []
    for t in range(19):
        url= "https://pokeapi.co/api/v2/type/"+str(t+1)+"/"
        u=t+1
        payload = {}
        headers = {}
        response = requests.request("GET", url)
        data=response.json()
        if u==1:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L1.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==2:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L2.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==3:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L3.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==4:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L4.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==5:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L5.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==6:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L6.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==7:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L7.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==8:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L8.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==9:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L9.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==10:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L10.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==11:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L11.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==12:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L12.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==13:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L13.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==14:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L14.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==15:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L15.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==16:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L16.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==17:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L17.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==18:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L18.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==19:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon'][i]['pokemon']['name'])
                    L19.append(ko)
                    i+1
            except IndexError:
                continue
        else:
            if u==10001:
                try:
                    for i in range(50000000000000000000000000000000):
                        ko=(data['pokemon'][i]['pokemon']['name'])
                        LR.append(ko)
                        i+1
                except IndexError:
                    break
            else:
                continue
            
                    
        
    x={'Normal':L1, 'Lucha':L2, 'Volador':L3, 'Veneno':L4, 'Tierra':L5, 'Roca':L6, 'Bicho':L7, 'Fantasma':L8, 'Acero':L9, 'Fuego':L10, 'Agua': L11, 'Planta': L12, 'Electricos': L13, 'Psíquico':L14, 'Hielo':L15, 'Dragón':L16, 'Siniestro':L17, 'Hada': L18,}
    return x
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
tip=tipos()
#LONGITUD
max_length = max(len(v) for v in tip.values())
for key, values in tip.items():
    while len(values) < max_length:
        values.append(0)
#WB
wb=Workbook()
wb.save('GrType.xlsx')
pat = "GrType.xlsx"  
wb = load_workbook(pat)
ws = wb.active
start_row = ws.max_row + 1
a=list(tip.keys())
ws.append(a)
start_row += 1

for row in zip(*tip.values()):
    ws.append(row)

wb.save(pat)
print('Ya')
