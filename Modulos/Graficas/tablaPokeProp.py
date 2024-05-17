import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Fill, Font
from openpyxl.styles import Alignment
from openpyxl.cell import Cell

#DANIEL CÁRDENAS ADAME


import os
import json
path=os.getcwd()
pathConsultas=os.path.join(path,'Consultas')
archivoFinal=os.path.join(pathConsultas,'Información.xlsx')


def nameToHex(name):
    if name=='negro':
        hexa='000000'
    if name=='azul':
        hexa='0000ff'
    if name=='amarillo':
        hexa='ffdd00'
    if name=='marrón':
        hexa='82401d'
    if name=='rojo':
        hexa='ff0000'
    if name=='blanco':
        hexa='ffffff'
    if name=='gris':
        hexa='888888'
    if name=='verde':
        hexa='00ff00'
    if name=='rosa':
        hexa='ff57e6'
    if name=='morado':
        hexa='b300ff'
    return hexa
    


def guardarPoke(nombre,pokeInfo,archivo=archivoFinal,hoja='Fornis'):
    try:
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivo):
            exists=1
            wb = load_workbook(archivo)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'hoja'
        if hoja in wb.sheetnames:
            ws = wb[hoja]
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet(hoja)
        
      
        #lkjlksjflasf
        #ahora si comenzar el proceso de hacer la tabla
        #FORMATO:
        #COLOR
        #TIPO
        #GRUPO
        #NUMERO
        #HABITAT
      
        i=1
        while True:
              cellValue= ws.cell(row=i, column=1).value
              cellValueNext= ws.cell(row=i+1, column=1).value
              if cellValue==None and cellValueNext==None:
                  fila=i+1
                  if i==1:
                      fila=1
                  break
              else:
                  i+=1


        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
        #título
        ws.cell(row=fila, column=1).value= nombre.capitalize()
        colorFill = PatternFill(start_color=nameToHex(pokeInfo['color']),
                   end_color=nameToHex(pokeInfo['color']),
                   fill_type='solid')
        ws.cell(row=fila, column=1).fill= colorFill
        if pokeInfo['color']=='negro':
            ws.cell(row=fila, column=1).font = Font(color='ffffff', bold=True)
        else:
            ws.cell(row=fila, column=1).font = Font(bold=True)
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal='center')

        if len(pokeInfo['tipo'])==1 and len(pokeInfo['grupo'])==1:
            ws.merge_cells('A'+str(fila)+':B'+str(fila))
        else:
            ws.merge_cells('A'+str(fila)+':C'+str(fila))


        colorFill = PatternFill(start_color='cccccc',
                   end_color='cccccc',
                   fill_type='solid')

        
        #color
        ws.cell(row=fila+1, column=1).value= 'Color'
        ws.cell(row=fila+1, column=1).font = Font(bold=True)
        ws.cell(row=fila+1, column=2).value= pokeInfo['color'].capitalize()


        #tipo
        if len(pokeInfo['tipo'])==1:
            ws.cell(row=fila+2, column=1).value= 'Tipo'
            ws.cell(row=fila+2, column=1).font = Font(bold=True)
            ws.cell(row=fila+2, column=2).value= pokeInfo['tipo'][0].capitalize()

        elif len(pokeInfo['tipo'])==2:
            ws.cell(row=fila+2, column=1).value= 'Tipos'
            ws.cell(row=fila+2, column=1).font = Font(bold=True)
            ws.cell(row=fila+2, column=2).value= pokeInfo['tipo'][0].capitalize()
            ws.cell(row=fila+2, column=3).value= pokeInfo['tipo'][1].capitalize()
        


        #grupo
        if len(pokeInfo['grupo'])==1:
            ws.cell(row=fila+3, column=1).value= 'Grupo'
            ws.cell(row=fila+3, column=1).font = Font(bold=True)
            ws.cell(row=fila+3, column=2).value= pokeInfo['grupo'][0].capitalize()

        elif len(pokeInfo['grupo'])==2:
            ws.cell(row=fila+3, column=1).value= 'Grupos'
            ws.cell(row=fila+3, column=1).font = Font(bold=True)
            ws.cell(row=fila+3, column=2).value= pokeInfo['grupo'][0].capitalize()
            ws.cell(row=fila+3, column=3).value= pokeInfo['grupo'][1].capitalize()

        #numero
        ws.cell(row=fila+4, column=1).value= 'No. Nacional'
        ws.cell(row=fila+4, column=1).font = Font(bold=True)
        ws.cell(row=fila+4, column=2).value= pokeInfo['id']


        #Hábitat
        ws.cell(row=fila+5, column=1).value= 'Hábitat'
        ws.cell(row=fila+5, column=1).font = Font(bold=True)
        ws.cell(row=fila+5, column=2).value= pokeInfo['habitat'].capitalize()


        for i in range(1,6):
            ws.cell(row=fila+i, column=1).fill=colorFill
        
        wb.save(archivo)
        return True

    except:
        return False
    
    
  
  
