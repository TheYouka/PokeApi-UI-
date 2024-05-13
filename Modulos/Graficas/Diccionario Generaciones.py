def gen():
    L1, L2, L3, L4, L5, L6, L7, L8, L9= [], [], [], [], [], [], [], [], []
    for t in range(9):
        url= "https://pokeapi.co/api/v2/generation/"+str(t+1)+"/"
        u=t+1
        payload = {}
        headers = {}
        response = requests.request("GET", url)
        data=response.json()
        if u==1:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L1.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==2:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L2.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==3:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L3.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==4:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L4.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==5:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L5.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==6:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L6.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==7:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L7.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==8:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L8.append(ko)
                    i+1
            except IndexError:
                continue
        elif u==9:
            try:
                for i in range(50000000000000000000000000000000):
                    ko=(data['pokemon_species'][i]['name'])
                    L9.append(ko)
                    i+1
            except IndexError:
                continue
            
        else:
            break
    x={'Generación 1': L1, 'Generación 2': L2, 'Generación 3': L3, 'Generación 4': L4, 'Generación 5': L5, 'Generación 6': L6, 'Generación 7':L7, 'Generación 8': L8, 'Generación 9': L9}
    return x       
import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
dicg=gen()
#LONGITUD
max_length = max(len(v) for v in dicg.values())
for key, values in dicg.items():
    while len(values) < max_length:
        values.append(0)
#WB
wb=Workbook()
wb.save('GrGn.xlsx')
pat = "GrGn.xlsx"  
wb = load_workbook(pat)
ws = wb.active
start_row = ws.max_row + 1
a=list(dicg.keys())
ws.append(a)
start_row += 1

for row in zip(*dicg.values()):
    ws.append(row)

wb.save(pat)
print('Ya')

