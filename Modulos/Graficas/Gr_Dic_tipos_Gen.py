from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import os
from datetime import datetime
path=os.getcwd()

from Modulos import pokeOffline as off

#ESTE PROGRAMA SOLO SE DEBE DE CORRER DESDE MAIN.PY
#Daniel Cárdenas Adame - Añadió funcionalidad de revisar por generación


#aquí se está suponiendo que se corre desde main.py
def grafTipos(gen, fila=1, archivo='Gráficas.xlsx'):
    titulo="Cantidad de Pokémon de cada tipo en la generación "+str(gen)
    path=os.getcwd()
    pathConsultas=os.path.join(path,'Consultas')
    pathGrafPng=os.path.join(pathConsultas,'Gráficas')
    archivoFinal=os.path.join(pathConsultas,archivo)
    
    try:
        
        #Grafica de pastel de tipos
        tipoDic=off.abrirRegistro('Tipos.txt')
        genList=off.abrirRegistro('Gen'+str(gen)+'.txt')
        #sacando la cantidad de pokemon en cada uno
        tipoLen={}
        for tipo in list(tipoDic.keys()):
            tipoLen[tipo]=0
        for tipo in list(tipoDic.keys()):
            for name in genList:
                if name in tipoDic[tipo]:
                    tipoLen[tipo]+=1
        #se va contando la cantidad de pokemones de cada color y se le suma a colorLen
                    
        #se puede sacar la longitud de cada uno después


        #se saca total de pokemones en la generacion
        Tot=0
        for i in list(tipoLen.keys()):
            Tot+=tipoLen[i]
        #ya se sacó el total

        T=[]
        porcentajes=[]
        for tipo in list(tipoDic.keys()):
            if tipoLen[tipo]!=0:
                T.append(tipo.capitalize())
                p=(tipoLen[tipo]*100)/Tot
                porcentajes.append(p)


        #Crear gráfica
        plt.pie(porcentajes, labels=T, autopct="%1.1f%%", startangle=90)
        plt.axis("equal")
        plt.title(titulo)

        current_datetime = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
         #Guardar la gráfica como imagen
        image_path = os.path.join(pathGrafPng, 'grafica-'+current_datetime+'.png')
        plt.savefig(image_path)

        #se muestra la gráfica al usuario
        plt.show()

        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Tipos'
        if 'Tipos' in wb.sheetnames:
            ws = wb['Tipos']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Tipos')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)
        

        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False
