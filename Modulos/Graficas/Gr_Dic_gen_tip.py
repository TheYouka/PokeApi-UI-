def tipos():
    # Crear una lista de listas para almacenar datos
    listas = [[] for _ in range(18)]
    path=os.path.join(os.path.dirname(__file__), 'GrType.xlsx')
    wb=load_workbook(filename=path)
    ws=wb.active
    s=-1
    for i in range (1,19):
        s=s+1
        for row in ws.iter_rows(min_row=2, max_col=i, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    listas[s].append(valor)
                    continue
    tipos_nombres = [
        'L1', 'L2', 'L3', 'L4', 'L5', 'L6', 
        'L7', 'L8', 'L9', 'L10', 'L11', 'L12', 
        'L13', 'L14', 'L15', 'L16', 'L17', 'L18'
    ]
    dic_tipos={}
    for i in range(len(tipos_nombres)):
        dic_tipos[tipos_nombres[i]] = listas[i]

    y=dic_tipos
    wb.save(filename=path)
    return(y)
def gen():
    path=os.path.join(os.path.dirname(__file__), 'GrGn.xlsx')
    wb =load_workbook(filename=path)
    ws = wb.active
    s=-1
    listas={}
    c=0
    for i in range(1,8):
        c=c+1
        if i==2:
            i=9
        elif i==3:
            break
        elif s!=8:
            s=s+1
        else:
            pass
        nam=f"L{c}"
        listas[nam]=[]
        for row in ws.iter_rows(min_row=2, max_col=i, values_only=True):
            valor = row[s]
            if valor is None:
                continue
                
            if valor != 0:
                listas[nam].append(valor)
                continue
        s=8
    x=listas
    wb.save(filename=path)
    return(x)

import requests
from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import os
#Grafica de comparacion de la gen 1 y 9 en cantidad de tipos
dtip=tipos()
dgen=gen()
#crear conjuntos e intersecciones
set_gen=set()
a=dgen['L1']
set_gen.update(a)
dicgyt1={}
for n in range (1,19):
    b=dtip['L'+str(n)]
    nam=f"L{n}"
    dicgyt1[nam]=[]
    set_tip=set()
    set_tip.update(b)
    c=set_gen & set_tip
    dicgyt1[nam].append(c)
    if c==set():
        dicgyt1[nam]='0'
    else:
        pass
        
#Crear las int. para gen 9
set_genf=set()
a=dgen['L2']
set_genf.update(a)
dicgyt2={}
for n in range(1,19):
    b=dtip['L'+str(n)]
    nom=f"L{n}"
    dicgyt2[nom]=[]
    set_tip2=set()
    set_tip2.update(b)
    c=set_genf & set_tip2
    dicgyt2[nom].append(c)
    if c==set():
        dicgyt2[nom]='0'
    else:
        pass
#Encontrar la cantidad de cada tipo
ln9=[]
for conjunto_lista in dicgyt2.values():
    for conjunto in conjunto_lista:
        if conjunto=='0':
            ln9.append(0)
        else:
            ln9.append(len(conjunto))
    
ln1=[]
for conj in dicgyt1.values():
    for lo in conj:
        if lo=='0':
            ln1.append(0)
        else:
            ln1.append(len(lo))
#hacer gráfica
etiquetas=['Normal', 'Lucha', 'Volador', 'Veneno', 'Tierra', 'Roca', 'Bicho', 'Fantasma', 'Acero', 'Fuego', 'Agua', 'Planta', 'Electricos', 'Psíquico', 'Hielo', 'Dragón', 'Siniestro', 'Hada']
#nuestros valores son ln1 y ln9
co=np.arange(len(ln1))
an=.40
fig, ax= plt.subplots()
ax.bar(co-an/2, ln1, an, label='Generación 1', color='#3B7AD4')
ax.bar(co+an/2, ln9, an, label='Generación 9', color='#D4C739')

for i,j in zip(co, ln1):
    ax.annotate(j, xy=(i -0.4, j +0.2))
for i,j in zip(co, ln9):
    ax.annotate(j, xy=(i +0.1, j +0.2))
    
ax.set_title("Tipos de Pokemones por generación")
ax.set_ylabel("Cantidad de Pokemones")
ax.set_xticks(co)
ax.set_xticklabels(etiquetas, rotation=90)
plt.legend()

plt.show()
return
