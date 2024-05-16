from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
import os
import requests as req

#DANIEL CÁRDENAS ADAME

def getStats(pokemonID):
    statData=[0,0,0,0,0,0]
    # HP,  ATK,  DEF,  SPATK,  SPDEF,  SPE
    pkmnData=req.get('https://pokeapi.co/api/v2/pokemon/'+str(pokemonID)+'/').json()
    
    for i in range(6):
        statData[i]=pkmnData['stats'][i]['base_stat']

    return statData
    

def graphStats(statData,name,fila=1,archivo='Gráficas.xlsx'):

    path=os.getcwd()
    pathConsultas=os.path.join(path,'Consultas')
    archivoFinal=os.path.join(pathConsultas,archivo)
    
    try:
        
        names=['HP','Ataque','Defensa','At. Especial','Def. Especial','Velocidad']
        colorVal=[0,0,0,0,0,0]
        for i in range(6):
            colorVal[i]=statData[i]/255
        plt.rcParams['figure.figsize']=(9.2, 4.8)
        statColor= LinearSegmentedColormap.from_list('smogonType',[(0,'#ff0000'),(0.29,'#c5fc00'),(0.78,'#00ff00'),(1,'#00ff00')],N=256)
        fig, ej =plt.subplots()
        barras= ej.bar(names,statData, color=statColor(statData))
        plt.title('Habilidades de '+name)
        for rect, label in zip(barras, statData):
            height = rect.get_height()
            ej.text(rect.get_x() + rect.get_width() / 2, height, label)

         #Guardar la gráfica como imagen (esto es temporal)
        image_path = os.path.join(pathConsultas, 'grafica.png')
        plt.savefig(image_path, bbox_inches='tight')

        #Se muestra
        plt.show()

        
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Color'
        if 'Habilidades' in wb.sheetnames:
            ws = wb['Habilidades']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Habilidades')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)
        
        #Eliminar la imagen temporal
        os.remove(image_path)

        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False
