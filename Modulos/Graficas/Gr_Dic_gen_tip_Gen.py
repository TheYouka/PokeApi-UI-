from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import os
import numpy as np
path=os.getcwd()

if path[-33:-17]=='PokeApi-UI--main':
    import pokeOffline as off
elif path[-16:]=='PokeApi-UI--main':
    from Modulos import pokeOffline as off

def compararTipos(gen1=1,gen2=9,fila=1,archivo='Gráficas.xlsx'):

    path=os.getcwd()
    pathConsultas=os.path.join(path,'Consultas')
    archivoFinal=os.path.join(pathConsultas,archivo)
    tipoDic=off.abrirRegistro('Tipos.txt')
    genList1=off.abrirRegistro('Gen'+str(gen1)+'.txt')
    genList2=off.abrirRegistro('Gen'+str(gen2)+'.txt')

    try:

        tipos=etiquetas=['normal', 'lucha', 'volador', 'veneno', 'tierra', 'roca', 'bicho', 'fantasma', 'acero', 'fuego', 'agua', 'planta', 'eléctrico', 'psíquico', 'hielo', 'dragón', 'siniestro', 'hada']
        
        #Encontrar la cantidad de cada tipo de cada generación
        listaGen1=[]
        for tipoName in tipos:
            tipoSet=set(tipoDic[tipoName])
            genSet=set(genList1)
            pokeInTipoInGen=tipoSet.intersection(genSet)
            listaPokes=list(pokeInTipoInGen)
            listaGen1.append(len(listaPokes))
            



        listaGen2=[]
        for tipoName in tipos:
            tipoSet=set(tipoDic[tipoName])
            genSet=set(genList2)
            pokeInTipoInGen=tipoSet.intersection(genSet)
            listaPokes=list(pokeInTipoInGen)
            listaGen2.append(len(listaPokes))
        
        
        #hacer gráfica
        etiquetas=['Normal', 'Lucha', 'Volador', 'Veneno', 'Tierra', 'Roca', 'Bicho', 'Fantasma', 'Acero', 'Fuego', 'Agua', 'Planta', 'Eléctrico', 'Psíquico', 'Hielo', 'Dragón', 'Siniestro', 'Hada']
        plt.rcParams['figure.figsize']=(11.15, 4.8)
        
        
        
        co=np.arange(len(listaGen1))
        an=.40
        fig, ax= plt.subplots()
        ax.bar(co-an/2, listaGen1, an, label='Generación '+str(gen1), color='#3B7AD4')
        ax.bar(co+an/2, listaGen2, an, label='Generación '+str(gen2), color='#D4C739')

        for i,j in zip(co, listaGen1):
            ax.annotate(j, xy=(i -0.4, j +0.2))
        for i,j in zip(co, listaGen2):
            ax.annotate(j, xy=(i +0.1, j +0.2))
            
        ax.set_title("Tipos de Pokémon por generación")
        ax.set_ylabel("Cantidad de Pokémon")
        ax.set_xticks(co)
        ax.set_xticklabels(etiquetas, rotation=90)
        plt.legend()

        
         #Guardar la gráfica como imagen (esto es temporal)
        image_path = os.path.join(pathConsultas, 'grafica.png')
        plt.savefig(image_path, bbox_inches='tight')

        #Se muestra
        plt.show()

        
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Color'
        if 'Comparaciones de tipos' in wb.sheetnames:
            ws = wb['Comparaciones de tipos']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Comparaciones de tipos')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)
        
        #Eliminar la imagen temporal
        os.remove(image_path)

        
        return True


    except Exception as e:
        print(f"Error: {e}")
        return False


        
