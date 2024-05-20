from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from datetime import datetime
import os
path=os.getcwd()

from Modulos import pokeOffline as off

#ESTE PROGRAMA SOLO SE DEBE DE CORRER DESDE MAIN.PY
#Daniel Cárdenas Adame - Añadió funcionalidad de revisar por generación


#aquí se está suponiendo que se corre desde main.py
def grafColor(gen, fila=1, archivo='Gráficas.xlsx'):
    titulo="Cantidad de Pokémon de cada color en la generación "+str(gen)
    path=os.getcwd()
    pathConsultas=os.path.join(path,'Consultas')
    pathGrafPng==os.path.join(pathConsultas,'Gráficas')
    archivoFinal=os.path.join(pathConsultas,archivo)
    try:
        
        #Grafica de pastel de colores
        colorDic=off.abrirRegistro('Color.txt')
        genList=off.abrirRegistro('Gen'+str(gen)+'.txt')
        #sacando la cantidad de pokemon en cada uno
        colorLen={}
        for color in list(colorDic.keys()):
            colorLen[color]=0
        for color in list(colorDic.keys()):
            for name in genList:
                if name in colorDic[color]:
                    colorLen[color]+=1
        #se va contando la cantidad de pokemones de cada color y se le suma a colorLen
                    
        #se puede sacar la longitud de cada uno después


        #se saca total de pokemones en la generacion
        Tot=len(genList)
        #ya se sacó el total

        T=[]
        porcentajes=[]
        for color in list(colorDic.keys()):
            T.append(color.capitalize())
            if colorLen[color]!=0:
                p=(colorLen[color]*100)/Tot
            porcentajes.append(p) 


        #Crear gráfica
        colores=['beige','blue','brown','grey','green','pink','purple','red','white', 'yellow']
        plt.pie(porcentajes, labels=T, autopct="%1.1f%%", startangle=90, colors=colores)
        plt.axis("equal")
        plt.title(titulo)

        current_datetime = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
         #Guardar la gráfica como imagen
        image_path = os.path.join(pathGrafPng, 'grafica-'+current_datetime+'.png')
        plt.savefig(image_path)

        #Se muestra
        plt.show()

        
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Color'
        if 'Color' in wb.sheetnames:
            ws = wb['Color']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Color')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)
        
        

        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False

