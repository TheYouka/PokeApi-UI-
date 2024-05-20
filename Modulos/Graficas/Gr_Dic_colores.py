from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
from datetime import datetime
import os

#ESTE PROGRAMA SOLO SE DEBE DE CORRER DESDE MAIN.PY

def grafColor(fila=1, archivo='Gráficas.xlsx'):

    
    path=os.getcwd()
    pathConsultas=os.path.join(path,'Consultas')
    pathModulos=os.path.join(path,'Modulos')
    pathGraficas=os.path.join(pathModulos,'Graficas')
    pathGrafPng=os.path.join(pathConsultas,'Gráficas')
    archivoFinal=os.path.join(pathConsultas,archivo)
    
    try:

        #Grafica de pastel de colores
        wb =load_workbook(os.path.join(pathGraficas,"GrColor.xlsx"))
        ws = wb.active
        #1r for para manipular las columnas y filas, cree el dic de listas para almacenar listas con nombres de variables
        #2do for para manipular el excel y no contar 0s
        s=-1
        listas={}
        for i in range(0,10):
            c=i+1
            s=s+1
            nam=f"L{c}"
            listas[nam]=[]
            for row in ws.iter_rows(min_row=2, max_col=c, values_only=True):
                valor = row[s]
                if valor is None:
                    continue
                    
                if valor != 0:
                    listas[nam].append(valor)
                    continue
        #Sacar la long de c/u
        dic=[]
        for i in range (0,10):
            d=len(listas['L'+str(i+1)])
            dic.append(d)
        #sacar total y porcentajes de c/u
        total=sum(dic)
        porcentajes=[]
        for i in range (10):
            a=dic[i]
            p=(a*100)/1025
            porcentajes.append(p) 
        #Para asignar los titulos en una lista
        T=[]
        for i in range (1,11):
            t=ws.cell(row=1, column=i)
            T.append(t.value)
        #Crear gráfica
        colores=['beige','blue','brown','grey','green','pink','purple','red','white', 'yellow']
        plt.pie(porcentajes, labels=T, autopct="%1.1f%%", startangle=90, colors=colores)
        plt.axis("equal")
        plt.title("Cantidad de Pokémon de cada color")
        wb.save(os.path.join(pathGraficas,"Grcolor.xlsx"))

        
        current_datetime = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
         #Guardar la gráfica como imagen
        image_path = os.path.join(pathGrafPng, 'grafica-'+current_datetime+'.png')
        plt.savefig(image_path)

        
        plt.show()
        
        #Cargar o crear el archivo Excel
        if os.path.exists(archivoFinal):
            exists=1
            wb = load_workbook(archivoFinal)
        else:
            wb = Workbook()
            exists=0
        
        #Seleccionar o crear la hoja 'Color'
        if 'Color' in wb.sheetnames:
            ws = wb['Color']
        else:
            if exists!=1:
                for sheet in wb.sheetnames:
                    del wb[sheet]
            ws = wb.create_sheet('Color')
        
        #Insertar la imagen en la fila especificada
        img = Image(image_path)
        ws.add_image(img, 'A'+str(fila))
        
        #Guardar el archivo Excel
        wb.save(archivoFinal)

        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False
