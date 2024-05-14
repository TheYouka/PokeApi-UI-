import requests
import json
from openpyxl import load_workbook
from openpyxl import Workbook
#crear wb
#wb=Workbook()
#wb.save('Grupos de huevos.xlsx')
dgh={}
for i in range (1, 15):
    url="https://pokeapi.co/api/v2/egg-group/"+str(i)+"/"
    response = requests.request("GET", url)
    data=response.json()
    k=(data['name'])
    nam=k
    dgh[nam]=[]
    j=0
    try:
        for j in range(len(data['pokemon_species'])):
            ko=(data['pokemon_species'][j]['name'])
            dgh[nam].append(ko)
            j+1
    except IndexError:
        pass
values=[]
#cambiar longitud antes de agregarlo
max_length = max(len(v) for v in dgh.values())
for key, values in dgh.items():
    while len(values) < max_length:
        values.append(0)
#ESCRIBIENDO EN EL WB
pat = 'Gpohvs.xlsx'  
wb = load_workbook(pat)
ws = wb.active
start_row = ws.max_row + 1
a=list(dgh.keys())
ws.append(a)
start_row += 1
for row in zip(*dgh.values()):
    ws.append(row)

wb.save(pat)
print('Se ha guardado en el excel '+str(pat))