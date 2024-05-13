from openpyxl import load_workbook
from openpyxl import Workbook
import matplotlib.pyplot as plt
wb =load_workbook("GrType.xlsx")
ws = wb.active
s=-1
listas={}
for i in range(0,18):
    c=i+1
    s=s+1
    nam=f"L{c}"
    listas[nam]=[]
    for row in ws.iter_rows(min_row=2, max_col=c, values_only=True):
        valor = row[s]
        if valor is None:
            continue
            
        if valor != 0:
            listas[nam].append(valor)
            continue
#Sacar la long de c/u
tl=[]
for i in range (0,18):
    d=len(listas['L'+str(i+1)])
    tl.append(d)
#sacar total y porcentajes de c/u
total=sum(tl)
porcentajes_tip=[]
for i in range (18):
    a=tl[i]
    p=(a*100)/1025
    porcentajes_tip.append(p)
#Para asignar los titulos en una lista
titulos=[]
for i in range (1,19):
    t=ws.cell(row=1, column=i)
    titulos.append(t.value)

#Crear gráfica
plt.pie(porcentajes_tip, labels=titulos, autopct="%1.1f%%", startangle=90)
plt.axis("equal")
plt.title("Cantidad de Pokemones de cada tipo")
plt.show()
wb.save("GrType.xlsx")